# -*- coding: utf-8 -*-
"""
Created on Sun Mar 29 14:12:56 2020

@author: Matthew
"""

#Continuing Automate The Boring Stuff
#Chapter 4 - Lists
spam = ['cat', 'bat', 'duck']
spam[1]
spam[0]
#lists in lists
spam2 = [['cat', 'bat', 'duck'], [1, 2, 3, 4]]
spam2[1][2]
#negative indices start from the last
spam2[0][-1]
#slices of lists
spam2[0][0:2]
# 'up to but not including the last number' - ex. not 'duck'
len(spam2[0])
#Assigning
spam2[0][1] = 'potato'
#can overwrite entire list if syntax is off
spam2[0] = 'gone'
print(spam2)
#deletions
del spam2[1][2]
#Ex program
catNames = []
while True:
    print('Enter the name of cat' + str(len(catNames)+1)+
          ' (Or enter nothing to stop.):')
    name = input()
    if name =='':
        break
    catNames = catNames + [name] #list concatenation
print('The cat names are:')
for name in catNames:
    print(' '+ name)
# loops and lists
supplies = ['pens', 'staplers', 'flame-throwers', 'binders']
for i in range(len(supplies)):
    print('Index' + str(i) + 'in supplies is : ' + supplies[i])
#in and not in
print(spam2)
'cat' in spam2[0] #True
'dog' in spam2[0] #False
#example program
myPets = ['Zophie', 'Pooka', 'Fat-tail']
print('Enter a pet name:')
name = input()
if name not in myPets:
    print('I do not have a pet named ' + name)
else:
    print(name + ' is my pet.')
#Multiple assignment trick - thb dont understand this one.
cat = ['fat', 'black', 'loud']
size, color, disposition = cat
#Methods - similar to function but it will be 'called on' a value.
spam = ['hello', 'hi', 'howdy', 'heyas']
spam.index('heyas')
#if duplicates, displays the first position in a list
spam = ['hello', 'hi', 'hello', 'howdy', 'heyas']
spam.index('hello')
# .append() - on the end
spam.append('potato')
#insert(#position, list)
spam.insert(2, 'oregano')
#both insert and append are only applicable on lists []
spam.remove('hello')
#also just removes the 1st instance of a value
# del() if you know the position you want to remove, remove() if you know the value
#sort() integers and flaots low to high, strings alphabetical 
spam.sort()
spam
#alternatively
spam.sort(reverse=True)
spam
#is case sensitive, all caps will be sorted, followed by all lowercase
#simplified 8 ball program
import random
messages = ['It is certain',
    'It is decidedly so',
    'Yes definitely',
    'Reply hazy try again',
    'Ask again later',
    'Concentrate and ask again',
    'My reply is no',
    'Outlook not so good',
    'Very doubtful']
print(messages[random.randint(0, len(messages) - 1)])
#strings are essentially lists at the end of the day, ex: 
name = 'Zophia'
name[1]
for i in name:
    print('***'+ i +'***')

#testing the updating of things on github across both computers
#tuples  are lists that are immutable - cant be modified - like strings
#Used with round brackets ()
eggs = ('hello', 42, 0.4)
#trailing comma indicates a tuple instead of a regular value
fish = (2,) #tuple
fish = (2) #integer value, equivalent to fish = 2 
#ordered sequence that dont change - use a tuple
#tuples and lists can be converted between each other
fish = (2,3,4,5)
fish2 = list((2,3,4,5))
fish2
#assigning a variable to a list is a reference to the list,
#not the actual list itself, EX:
spam = [1,2,3,4,5]
cheese = spam
cheese[1] = 'hello'
spam
# they both refer to the same list, not the case with tuples
#the variable assigned to a tuple IS the tuple
#Passing reference:
def eggs(someParameter):
    someParameter.append('Hello')
spam = [1,3,4]
eggs(spam)
print(spam)
#making true copys:
import copy
spam = ['A','B','C','D']
cheese = copy.copy(spam)
cheese[2] = 43
spam
cheese
#2 different lists like you would expect with other data types
#copy.deepcopy() will do the same with lists that contain internal lists

#practice problems ch4

def stringy(string):
    st = ''
    for i in range(len(string)):
        if i > 0:
            if i == len(string) -1:
                st = st + ' and '
            else:
                st = st + ', '
        st = st + string[i];
    return st
spam = ['apples', 'bananas', 'tofu', 'cats']
stringy(spam)
#problem 2
grid = [['.', '.', '.', '.', '.', '.'],
        ['.', 'O', 'O', '.', '.', '.'],
        ['O', 'O', 'O', 'O', '.', '.'],
        ['O', 'O', 'O', 'O', 'O', '.'],
        ['.', 'O', 'O', 'O', 'O', 'O'],
        ['O', 'O', 'O', 'O', 'O', '.'],
        ['O', 'O', 'O', 'O', '.', '.'],
        ['.', 'O', 'O', '.', '.', '.'],
        ['.', '.', '.', '.', '.', '.']]

def layer(row):
    for j in range(len(row[0])):
        for i in range(len(row)):
            print(row[i][j], end = '')
        print('')
layer(grid)
#end ch4

#start chapter 5 - Dictionaries and Structuring Data
#Dictionaries - indexed with keys > key-value pairs
myCat = {'size': 'fat', 'colour': 'grey', 'disposition': 'loud'}
#size, colour, disposition are the keys, all associated with their values
myCat['size']
'My cat has ' + myCat['colour'] + ' fur.'
#can use integers as keys as well. Can be any number
spam = {12345: 'Luggage Combination', 42: 'The Answer'}
#dictionaries are not ordered like lists or tuples
#therefore cant be sliced
birthdays = {'Alice': 'Apr 1', 'Bob': 'Dec 12', 'Carol': 'Mar 4'}
while True:
    print('Enter a name: (Blank to quit)')
    name = input()
    if name == '':
        break
    if name in birthdays:
        print(birthdays[name] + ' is the birthday of ' +name)
    else:
        print('I do not have birthday information for ' + name)
        print('What is their birthday?')
        bday = input()
        birthdays[name] = bday
        print('Birthday database updated')
#this data does not get saved permanently in the dictionary - will be taught later in the book
#keys(), values(), items()
spam = {'colour': 'red', 'age': 42}
for v in spam.values():
    print(v)
for k in spam.items():
    print(k)
#more multiple assignment things
for k, v in spam.items():
    print('Key: '+ k + ' Value: ' + str(v))
#can still use in and not in to check for stuff in the dictionary
spam = {'name': 'Zophie', 'age': 7}
'name' in spam.keys()
'color' not in spam.keys()
'Zophie' in spam.values()
#get()
picnicItems = {'apples': 5, 'cups': 2} 
#inserting a default fallback value if the key is missing (0 in this case)
'I am bringing ' + str(picnicItems.get('cups', 0)) + ' cups.'
'I am bringing ' + str(picnicItems.get('eggs', 0)) + ' eggs.'
#setdefault() - only if that key doesnt have a value
spam = {'name': 'Pooka', 'Age': 5}
if 'colour' not in spam:
    spam['colour'] = 'black'
#setdefault condenses this down to 1 line
spam = {'name': 'Pooka', 'Age': 5}
spam.setdefault('colour', 'black')
spam
#setdefault ensures that a key exists
message = 'It was a bright cold day in April, and the clocks were striking thirteen.'
count = {}
for character in message:
    count.setdefault(character, 0)
    count[character] = count[character] + 1
print(count)
#Pretty Print with pprint() and pformat()
import pprint
message = 'It was a bright cold day in April, and the clocks were striking thirteen.'
count = {}
for character in message:
    count.setdefault(character, 0)
    count[character] = count[character] + 1
pprint.pprint(count)
#sorts keys, adds newlines, way cleaner
#pprint.pformat() makes it into a string
print(pprint.pformat(count))
#modeling tic tac toe with a dictionary
theBoard = {'top-L': ' ', 'top-M': ' ', 'top-R': ' ',
            'mid-L': ' ', 'mid-M': ' ', 'mid-R': ' ',
            'low-L': ' ', 'low-M': ' ', 'low-R': ' '}
def printBoard(board):
    print(board['top-L'] + '|' + board['top-M'] + '|' + board['top-R'])
    print('-+-+-')
    print(board['mid-L'] + '|' + board['mid-M'] + '|' + board['mid-R'])
    print('-+-+-')
    print(board['low-L'] + '|' + board['low-M'] + '|' + board['low-R'])
printBoard(theBoard)
turn = 'X'
for i in range(9):
    printBoard(theBoard)
    print('Turn for ' + turn + '. Move on which space?')
    move = input()
    theBoard[move] = turn
    if turn == 'X':
        turn = 'O'
    else:
        turn = 'X'
printBoard(theBoard)
#simple dictionary pairings between single places and a state
#can do nested dictionaries for more complicated situations
allGuests = {'Alice': {'apples': 5, 'Pretzels':12},
             'Bob': {'ham sandwiches': 3, 'apples': 2},
             'Carol': {'cups': 3, 'apple pies': 1}}
def totalBrought(guests, item):
    numBrought = 0
    for k, v in guests.items():
        numBrought = numBrought + v.get(item, 0)
    return numBrought

print('Number of things being brought:')
print(' - Apples       ' + str(totalBrought(allGuests, 'apples')))
print(' - Cups         ' + str(totalBrought(allGuests, 'cups')))
print(' - Cakes        ' + str(totalBrought(allGuests, 'cakes')))
print(' - Ham Sandwiches ' + str(totalBrought(allGuests, 'ham sandwiches')))
print(' - Apple Pies ' + str(totalBrought(allGuests, 'apple pies')))
#dumb for this purpose, but can be extended to thousands of entries
#end chapter 5, didnt do the practice problems, so much typing, little thinking

#Chapter 6 - Manipulating Strings
#manipulating strings
spam = 'Say hi to Bob\'s mother' 
#escape character \ to allow the '  without ending the string
#   Escape character Prints as
#   \' Single quote
#   \" Double quote
#   \t Tab
#   \n Newline (line break)
#   \\ Backslash
print("Hello there! \nHow are you?\nI\'m doing fine.")
#raw strings (r...) completely ignore all escape characters
print(r'That is Carol\'s cat.')
#triple quotes for multiline strings
print('''Dear Alice,
      
Eve's cat has been arrested for catnapping, cat burglary, and extortion.

Sincerely,
Bob''')
# includes the newlines as part of the string, alt:
print('Dear Alice,\n\nEve\'s cat has been arrested for catnapping, cat burglary, and extortion.\n\nSincerely,\nBob')
#useful string methods
spam = 'Hello world!'
spam = spam.upper()
spam
spam = spam.lower()
spam
#these call new strings, not directly modifying the original
print('How are you?')
feeling = input()
if feeling.lower() == 'great':
    print('I feel great too!')
else:
    print('I hope the rest of your day is good.')
#putting input into lowercase allows you to accomodate case insensitivity
#isupper() and islower() allow you to test what a string is.
#will return a True/False value
#other isX() calls:
#isalpha() # is the string only letters?
#isalnum() # is the string only letters and numbers?
#isdecimal() # is the string only numbers?
while True:
    print('Enter your age:')
    age = input()
    if age.isdecimal():
        break
    print('Please enter a number for your age.')
while True:
    print('Select a new password (letters and numbers only):')
    password = input()
    if password.isalnum():
        break
    print('Passwords can only have letters and numbers.')
#startswith() and endswith() - self explanatory methods
# .join() and .split() switch strings and lists, ex:
'My name is Simon'.split()
' '.join(['My', 'name', 'is', 'Simon'])
'ABC'.join(['My', 'name', 'is', 'Simon'])
'MyABCnameABCisABCSimon'.split('ABC')
#Justifying text with - rjust(), ljust(), center()
'Hello'.rjust(10)
'Hello'.rjust(20)
'Hello World'.rjust(20,'*')
'Hello'.ljust(10, 'p')
#use things like this to make pretty outputs of data
def printPicnic(itemsDict, leftWidth, rightWidth):
    print('PICNIC ITEMS'.center(leftWidth + rightWidth, '-'))
    for k, v in itemsDict.items():
        print(k.ljust(leftWidth, '.') + str(v).rjust(rightWidth))
picnicItems = {'sandwiches': 4, 'apples': 12, 'cups': 4, 'cookies': 8000}
printPicnic(picnicItems, 12, 5)
printPicnic(picnicItems, 20, 6)
#Removing Whitespace with strip(), rstrip(), and lstrip()

#Table Printer Practice Project

tableData = [['apples', 'oranges', 'cherries', 'banana'],
['Alice', 'Bob', 'Carol', 'David'],
['dogs', 'cats', 'moose', 'goose']]

def printTable(doggo):
    colL = [0] * len(doggo) 
    for j in range(len(doggo)):
        for i in range(len(doggo[j])):
            if len(doggo[j][i]) > colL[j]:
                colL[i] = len(doggo[j][i])
    for y in range(len(doggo[0])):
        for x in range(len(doggo)):
            print(doggo[x][y].rjust(colL[x]), end = '')
        print('')

printTable(tableData)
#good enough

#Part 2 of the book: Automating Tasks

#Chapter 7
#Pattern matching with regular expressions
# used for searching for things that are in a consistent format,
#like phone numbers, emails, etc. 
#///
#Finding patterns without regular expressions
def isPhoneNumber(text):
    if len(text) !=12:
        return False
    for i in range(0, 3):
        if not text[i].isdecimal():
            return False
    if text[3] != '-':
        return False
    for i in range(4, 7):
        if not text[i].isdecimal():
            return False
    if text[7] != '-':
        return False
    for i in range(8, 12):
        if not text[i].isdecimal():
            return False
    return True
print(isPhoneNumber('495-635-5453'))     
print(isPhoneNumber('moshi moshi'))
#test
message = 'Call me at 415-555-1011 tomorrow. 415-555-9999 is my office.'
for i in range(len(message)):
    chunk = message[i:i+12]
    if isPhoneNumber(chunk):
        print('Phone number found: ' + chunk)
print('Done')
#now with regular expressions
import re
phoneNumRegex = re.compile(r'\d\d\d-\d\d\d-\d\d\d\d') # remember '(r' makes it a raw string
#regular expressions frequently use \ so passing raw strings save \\ each time
#.search() looks in a string it is passed for the regex patters
mo = phoneNumRegex.search('My number is 415-555-4242.')
print('Phone number found: ' + mo.group())
phoneNumRegex = re.compile(r'(\d\d\d)-(\d\d\d-\d\d\d\d)')
mo = phoneNumRegex.search('My number is 415-555-4242')
mo.group(1)
mo.group(2)
mo.group(0)
mo.group()
mo.groups()
areaCode, mainNumber = mo.groups()
print(areaCode)
print(mainNumber)
#if area code in ()
phoneNumRegex = re.compile(r'(\(\d\d\d\)) (\d\d\d-\d\d\d\d)')
mo = phoneNumRegex.search('My number is (415) 555-4242')
mo.group(1)
#matching multiple groups with the pipe |, essentially 'this' or 'this'
heroRegex = re.compile(r'Batman|Tina Fey')
mo1 = heroRegex.search('Batman and Tina Fey')
mo1.group()
#if a string has both, will always be the first of the 2 incidents
#specifying a prefix
batRegex = re.compile(r'Bat(man|mobile|copter|bat)')
mo = batRegex.search('Batmobile lost a wheel')
mo.group()
#optional matching with a ?
batRegex = re.compile(r'Bat(wo)?man')
mo1 = batRegex.search('The Adventures of Batman')
mo1.group()
mo2 = batRegex.search('The Adventures of Batwoman')
mo2.group()
#(wo)? means that part of the string is optional in the search
#relate to finding phone numbers with or without a area code
phoneRegex = re.compile(r'(\d\d\d-)?\d\d\d-\d\d\d\d')
mo1 = phoneRegex.search('My number is 415-555-4242')
mo1.group()
mo2 = phoneRegex.search('My number is 555-4242')
mo2.group()
# a * can be used as a optional OR multiple segment search. 0 or more
batRegex = re.compile(r'Bat(wo)*man')
mo1 = batRegex.search('The Adventures of Batwowowowoman')
mo1.group()
# a + will mean 1 or more, so not 0.
#match specific repetitions with {}
#(ha){3} = (ha)(ha)(ha)
#or (ha){3,5} - can match 3, 4, or 5 instances of (ha)
greedyHaRegex = re.compile(r'(Ha){3,5}')
mo1 = greedyHaRegex.search('HaHaHaHaHa')
mo1.group()
#greedy (longest string) vs nongreedy '?' (shortest string possible)
nongreedyHaRegex = re.compile(r'(Ha){3,5}?')
mo2 = nongreedyHaRegex.search('HaHaHaHaHa')
mo2.group()
#findall()  - return every match in a searched string
# will return list of strings if no groups, or a list of tuples 
#with their component strings if there are groups
phoneNumRegex = re.compile(r'\d\d\d-\d\d\d-\d\d\d\d')
phoneNumRegex.findall('Cell: 415-555-7788 work: 212-555-0000')
#groups -> tuples
phoneNumRegex = re.compile(r'(\d\d\d)-(\d\d\d)-(\d\d\d\d)')
phoneNumRegex.findall('Cell: 415-555-7788 work: 212-555-0000')

########################
#character classes
#\d Any numeric digit from 0 to 9.

#\D Any character that is not a numeric digit from 0 to 9.

#\w Any letter, numeric digit, or the underscore character.
#(Think of this as matching “word” characters.)

#\W Any character that is not a letter, numeric digit, or the
#underscore character.

#\s Any space, tab, or newline character. (Think of this as
#matching “space” characters.)

#\S Any character that is not a space, tab, or newline.
#########################

#more stuff i skipped - highlighted chart in book for referece
#case-insensitive matching - pass re.IGNORECASE or re.I as second argument
robocop = re.compile(r'robocop', re.I)
robocop.search('Robocop is part man, part machine, all cop.').group()
robocop.search('ROBOCOP protects the innocent.').group()
#substituting string with the sub() method argument
namesRegex = re.compile(r'Agent \w+')
namesRegex.sub('REDACTED', 'Agent Alice gave the documents to Agent Bob')

#end page 163

#regex can only handle 1 additional argument, can get multiple by piping

#Project - create phone number and email regex from clipboard
#skipping for now, get the idea and need to install pyperclip
phoneRegex = re.compile(r'''(
    (\d{3}|\(\d{3}\))?  `        # area code
    (\s|-|\.)?                   # separator
    \d{3}                        # first 3 digits
    (\s|-|\.)                    # separator
    \d{4}                        # last 4 digits
    (\s*(ext|x|ext.)\s*\d{2,5})? # extension
    )''', re.VERBOSE)

#skipping practice problems for now, regex's are getting dull af

#Chapter 8 - Reading and Writing files (on the harddrive)
#directories and filenames are not case-sensitive on windows.
import os
os.path.join('usr','bin','spam')
#remember that backslashes need to be escaped by another (therefore double)
myFiles = ['accounts.txt', 'details.csv', 'invite.docx']
for filename in myFiles:
    print(os.path.join('C:\\Users\\asweigart', filename))
#current working directory
import os
os.getcwd()
#absolute vs relative paths.
#sbsolute - always begins at root (ex. C:\), relative related to CWD
# .\ = this directory, ..\ parent directory
#creating directories
import os
os.makedirs('C:\\delicious\\walnut\\waffles')
#os.path Module
import os
os.path.abspath('.')
os.path.abspath('.\\Scripts')
os.path.isabs('.')
os.path.isabs(os.path.abspath('.'))
#?
path = 'C:\\Windows\\System32\\calc.exe'
os.path.basename(path)
os.path.dirname(path)
#or to get both together in a tuple
os.path.split(path)
os.path.getsize(path)
#There are three steps to reading or writing files in Python.
#1. Call the open() function to return a File object.
#2. Call the read() or write() method on the File object.
#3. Close the file by calling the close() method on the File object.
os.getcwd() #i dont get this
helloFile = open('C:\\users\\matthew\\hello.txt')
helloContent = helloFile.read()
helloContent
#to get a lsit of string values from the file, one for each line of text
sonnetFile = open('sonnet29.txt')
sonnetFile.readlines()
#Writing Files - writing mode with 'w' (overwrites), appending with 'a'
baconFile = open('bacon.txt', 'w')
baconFile.write('Hello World!\n')
baconFile.close()
baconFile = open('bacon.txt', 'a')
baconFile.write('Bacon is not a vegetable')
baconFile.close()
baconFile = open('bacon.txt')
content = baconFile.read()
baconFile.close()
print(content)
#with write() you have to add newlines to the ends manually with \n
#save variables with the shelve method
import shelve
shelfFile = shelve.open('mydata')
cats = ['Zophie','Pooka','Simon']
shelfFile['cats'] =cats
shelfFile.close()
#shelf files dont have to be opened in read/write mode, they can do both.
#Checking
shelfFile = shelve.open('mydata')
type(shelfFile)
shelfFile['cats']
shelfFile.close()
#like dictionaries, shelves have keys and values attached to them.
shelfFile = shelve.open('mydata')
list(shelfFile.keys())
list(shelfFile.values())
shelfFile.close()
#saving variables with the pprint.pformat() function
import pprint
cats = [{'name': 'Zophie', 'desc': 'chubby'}, {'name': 'Pooka', 'desc' : 'fluffy'}]
pprint.pformat(cats)
fileObj = open('myCats.py', 'w') 
fileObj.write('cats = ' + pprint.pformat(cats) + '\n')
fileObj.close()
#list of dictionaries stored in variable cats - retriveable even after
#shell is closed via pprint.pformat()
#when string is saved to a .py file, now a module that can be imported
import myCats
myCats.cats
myCats.cats[0]
myCats.cats[0]['name']

#PROJECT - Random quiz questions
import random
# The quiz data. Keys are states and values are their capitals.
capitals = {'Alabama': 'Montgomery', 'Alaska': 'Juneau', 'Arizona': 'Phoenix',
'Arkansas': 'Little Rock', 'California': 'Sacramento', 'Colorado': 'Denver',
'Connecticut': 'Hartford', 'Delaware': 'Dover', 'Florida': 'Tallahassee',
'Georgia': 'Atlanta', 'Hawaii': 'Honolulu', 'Idaho': 'Boise', 'Illinois':
'Springfield', 'Indiana': 'Indianapolis', 'Iowa': 'Des Moines', 'Kansas':
'Topeka', 'Kentucky': 'Frankfort', 'Louisiana': 'Baton Rouge', 'Maine':
'Augusta', 'Maryland': 'Annapolis', 'Massachusetts': 'Boston', 'Michigan':
'Lansing', 'Minnesota': 'Saint Paul', 'Mississippi': 'Jackson', 'Missouri':
'Jefferson City', 'Montana': 'Helena', 'Nebraska': 'Lincoln', 'Nevada':
'Carson City', 'New Hampshire': 'Concord', 'New Jersey': 'Trenton', 'New Mexico': 'Santa Fe', 'New York': 'Albany', 'North Carolina': 'Raleigh',
'North Dakota': 'Bismarck', 'Ohio': 'Columbus', 'Oklahoma': 'Oklahoma City',
'Oregon': 'Salem', 'Pennsylvania': 'Harrisburg', 'Rhode Island': 'Providence',
'South Carolina': 'Columbia', 'South Dakota': 'Pierre', 'Tennessee':
'Nashville', 'Texas': 'Austin', 'Utah': 'Salt Lake City', 'Vermont':
'Montpelier', 'Virginia': 'Richmond', 'Washington': 'Olympia', 'WestVirginia':
'Charleston', 'Wisconsin': 'Madison', 'Wyoming': 'Cheyenne'}
capitals
for quizNum in range(35):
    #creates quiz and answer key files
    quizFile = open('capitalsquiz%.txt' % (quizNum + 1), 'w') 
    answerKeyFile = open('capitalsquiz_answers%s.txt' % (quizNum +1), 'w')
    #write out header for this quiz
    quizFile.write('Name:\n\nDate:\n\nperiod:\n\n')
    quizFile.write((' ' * 20)+ 'State Capitals Quiz (Form %s)' % (quizNum +1))
    quizFile.write('\n\n')
    #shuffle order of the states
    states = list(capitals.keys())
    random.shuffle(states)
    #loop through all states, 1 question each
    for questionNum in range(50):
        #get right and wrong answers ( 4 total option)
        correctAnswer = capitals[states[questionNum]]
        wrongAnswers = list(capitals.values())
        del wrongAnswers[wrongAnswers.index(correctAnswer)]
        wrongAnswers = random.sample(wrongAnswers, 3)
        answerOptions = wrongAnswers + [correctAnswer]
        random.shuffle(answerOptions)
        #write questions and answer options to the quiz file
        quizFile.write('%s. What is the capital of %s?\n' % (questionNum + 1,
            states[questionNum]))
        for i in range(4):
            quizFile.write('    %s. %s\n' % ('ABCD'[1], answerOptions[1]))
        quizFile.write('\n')
        #write the answer key to a file
        answerKeyFile.write('%s. %s\n' % (questionNum + 1, 'ABCD'[
                answerOptions.index(correctAnswer)]))
        quizFile.close()
        answerKeyFile.close()

#i understand the code steps but i have no idea why this doesnt work
#practice programs skipped
#end chapter 8

#Start Chapter 9 - Organizing Files
import shutil, os
os.chdir('C:\\')
#moves a copy to the new directory with the same filename
shutil.copy('C:\\spam\\spam.txt', 'C:\\Delicious')
#moves a copy to the new directory with a new name
shutil.copy('C:\\spam\\eggs.txt', 'C:\\Delicous\\eggs2.txt')
#to move an entire tree, (From, to)
shutil.copytree('C:\\bacon', 'C\\Delicious')
#can move without copy tooo, shutil.move()
shutil.move('C:\\bacon.txt', 'C:\\eggs')
#careful with move, will overwrite existing files with the same name
#if you write the wrong name for a directory, will rename the file to that name
#move() is dumb and will do something no matter what you intended. 
### Deleting
os.unlink('path') # will delete the file at path
os.rmdir('path') # will delete empty folder at path
shutil.rmtree('path') # will remove the folder and all containing files
#PERMANENT DELETIONS <<< not recoverable. 
#good way to test for errors, comment out delete lines and print instead
#Ex:
import os
for filename in os.listdir():
    if filename.endswith('.rxt'):
        #os.unlink(filename)
        print(filename)
#this way will print, not delete the files due to TYPO!
#if working correctly, then can un-comment the os.unlink line
os.chdir('C:\\Users\Matthew\Documents\GitHub\BioInf')
os.getcwd()

import send2trash
baconFile = open('bacon.txt', 'a') #creates the file
baconFile.write('Bacon is not a vegetable.')
baconFile.close()
send2trash.send2trash('bacon.txt')
#os.walk() to walk down an entire directory, modifying all folders and files
os.getcwd()
import os
for folderName, subfolders, filenames in os.walk('C:\\Users\\Matthew\\Documents\\GitHub\\BioInf\\Delicious'):
    print('The current folder is ' + folderName)
    for subfolder in subfolders:
        print('SUBFOLDER OF ' + folderName + ': ' + subfolder)
        for filename in filenames:
            print('FILE INSIDE ' + folderName + ': ' + filename)
        print('')
#compressing files with the zipfile module
import zipfile, os
os.chdir('C:\\Users\\Matthew\\Documents\\GitHub\\BioInf')
exampleZip = zipfile.ZipFile('potato.zip')
exampleZip.namelist()
spamInfo = exampleZip.getinfo('spam.txt')
spamInfo.file_size
spamInfo.compress_size
'Compressed file is %sx smaller!' % (round(spamInfo.file_size / spamInfo.compress_size, 2))
exampleZip.close()
#extracting from Zip with extractall()
import zipfile, os
exampleZip = zipfile.ZipFile('potato.zip')
exampleZip.extractall()
exampleZip.close()
#alternatively, to a new folder
import zipfile, os
os.chdir('C:\\Users\\Matthew\\Documents\\GitHub\\BioInf')
exampleZip = zipfile.ZipFile('potato.zip')
exampleZip.extract('potato.zip', 'folderZ')
#done with all that shit, i never zip anyway
###################
#page 207
#Project Renaming dates from US format to EU format. 
import shutil, os, re
#creates regex to match files with the american date format
datePattern = re.compile(r"""^(.*?) # all text before the date
    ((0|1)?\d)-                     # one or 2 digets for the month
    ((0|1|2|3)?\d)-                 # one of 2 digets for the day
    ((19|20)\d\d)                   # four digets for the year
    (.*?)$                          # all text after the date
    """, re.VERBOSE)
#loop over the files in the working directory
for amerFilename in os.listdir('.'):
    mo = datePattern.search(amerFilename)
    #skip files without date
    if mo == None:
        continue
    #get the different parts of the filename
    beforePart = mo.group(1)
    monthPart = mo.group(2)
    dayPart = mo.group(4)
    yearPart = mo.group(6)
    afterPart = mo.group(8)

datePattern = re.compile(r"""^(1) # all text before the date
    (2 (3) )-       # one or two digets for the month                    
    (4 (5) )-       # one or two digets for the day
    (6 (7) )-       # four digets for the year
    (8)$            # all text after the date
    """, re.VERBOSE)
#form the new filename and rename the files
    #form euro filenames
euroFilename= beforePart + dayPart + '-' + monthPart + '-' + yearPart + afterPart
#get the full, absolute file paths
absWorkingDir = os.path.abspath('.')
amerFilename = os.path.join(absWorkingDir, amerFilename)
euroFilename = os.path.join(absWorkingDir, euroFilename)
#rename the files
print('Renaming "%s" to "%s"...' % (amerFilename, euroFilename))
#shutil.move(amerFilename, euroFilename) # uncomment after testing
#doesnt work becaue actually needs to run on test files to fill out variables
#############
#2nd program: Backing up to ZIP - page 210
#practice problems skipped for now. 
#End Chapter 9 


#Chapter 10 - Debugging - [lots to study and take notes on]
#Raising an Exception = “Stop running the code in this function
#and move the program execution to the except statement.”
raise Exception('This is an error message.') #Exception is case sensitive
#if no try / except statements, will just display the error message
#often code that calls a function that is built to handle and exception
def boxPrint(symbol, width, height):
    if len(symbol) !=1:
        raise Exception('Symbol must be a single character string.')
    if width <=2:
        raise Exception('Width must be greater than 2.')
    if height <=2:
        raise Exception('Height must be greater than 2.')
    print(symbol*width)
    for i in range(height -2):
        print(symbol + (' ' * (width - 2)) + symbol)
    print(symbol*width)
#This is now running it multiple times with different parameters
#'*' 4x4, '0' 20x5, etc.
for sym, w, h in(('*', 4, 4), ('0', 20, 5), ('x', 1, 3), ('ZZ', 3, 3)):
    try:
        boxPrint(sym, w, h)
    except Exception as err:
        print('An exception has happened: ' + str(err))
#try and except allows you to handle errors instead of a whole program crash
#Getting the traceback as a string:
def spam():
    bacon()
def bacon():
    raise Exception('This is an error messgage.')
spam()
#use traceback to get an error message for debugging without killin a running program
import traceback
try:
    raise Exception('This is an error message.')
except:
    errorFile = open('errorInfo.txt', 'w')
    errorFile.write(traceback.format_exc())
    errorFile.close()
    print('The traceback info was written to errorInfo.txt.')

import os
os.getcwd()
os.chdir('C:\\Users\\Matthew\\Documents\\GitHub\\BioInf')
os.getcwd()
#I dont understand how the working directory functions. Maybe due to
#the continued script ive been using for learning.
#############################
#Assertions: # Sanity check to see if code isnt doing something obvs wrong
# Components:
#• The assert keyword
#• A condition (that is, an expression that evaluates to True or False)
#• A comma
#• A string to display when the condition is False
podBayDoorStatus = 'open'
assert podBayDoorStatus == 'open', 'The pod bay doors need to be "open".'
podBayDoorStatus = 'I\'m sorry Dave. I\'m afraid I cant do that.'
assert podBayDoorStatus == 'open', 'The pod bay doors need to be "open".'
#assertions are error checking to make sure that something you are relying
#upon down the code isnt in a different state from something higher up.
#for programmer, not user errors

#Assertions can be disabled by passing -O when running python. 
#to speed up programs once debugging is done. 
##################
#Logging: #similar to prints to ensure a section has run. 
import logging
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
#Ex:
import logging
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
logging.debug('Start of program')

def factorial(n):
    logging.debug('Start of factorial(%s%%)' % (n))
    total = 1
    for i in range(n+1):
        total *= i 
        logging.debug('i is ' + str(i) + ', total is ' + str(total))
    logging.debug('End of factorial(%s%%)' % (n))
    return total

print(factorial(5))
logging.debug('End of program')
#Ex:
import logging
logging.basicConfig(level=logging.DEBUG, format=' %(asctime)s - %(levelname)s - %(message)s')
logging.debug('Start of program')

def factorial(n):
    logging.debug('Start of factorial(%s%%)' % (n))
    total = 1
    for i in range(1, n+1): #fixed this ex
        total *= i 
        logging.debug('i is ' + str(i) + ', total is ' + str(total))
    logging.debug('End of factorial(%s%%)' % (n))
    return total

print(factorial(5))
logging.debug('End of program')
#logging is useful because then print statements dont need to be all over your program to be disabled later.
#Can cancel all logging outputs with one disable statement

#logging can be at different levels depending on the severity
#debugging mode (toolbar) can be used to go line by line to evaluate what a program is doing,
#what every variable is set to at each line etc. 
#skipped for now - not relevant for the moment- also uses idle not spyder
#Skipped practice problems for same reason.
#end Chapter 10
#########################

#Chapter 11 -  Web Scraping
#using a program to download and process content from the web

#Project: mapIt.py with webbrowser module
import webbrowser
webbrowser.open('http://inventwithpython.com/')
#example of webbrowser - basically all it does
#now to make a program to open a map from a place on clipboard

#! python3
#mapIt.py - Launches a map in the browser using an address from command line

import webbrowser, sys, pyperclip
if len(sys.argv) >1:
    #get address from command line
    address = ' '.join(sys.argv[1:])
#webbrowser to launch web, sys to read command line stuff
#ensuring the command line arguments (>1) are all accounted for
else:
    #get address from clipboard
    address = pyperclip.paste()
webbrowser.open('https://www.google.com/maps/place/' + address)
#

#Downloading files from the web with Requests module
#requests.get()
import requests
res = requests.get('http://www.gutenberg.org/cache/epub/1112/pg1112.txt')
type(res)
#checking to see if the download has succeeded
res.status_code == requests.codes.ok
len(res.text)
#checking for errors - always call raise_for_status() after 
#calling requests.get() to make sure download worked b4 continueing
import requests
res = requests.get('https://inventwithpython.com/page_doesnt_exist')
try:
    res.raise_for_status()
except Exception as exc:
    print('There was a problem: %s' % (exc))
#saving downloaded files to the harddrive
#from here, can open() and write() file, but first need to open in 
    #'write binary' mode to keep encoding correct
import requests
res = requests.get('https://www.gutenberg.org/cache/epub/1112/pg1112.txt')
res.raise_for_status()
playFile = open('RomeoAndJuliet.txt', 'wb') #wb for write binary
for chunk in res.iter_content(100000):
    playFile.write(chunk)
playFile.close()
#To review, here’s the complete process for downloading and saving a file:
#1. Call requests.get() to download the file.
#2. Call open() with 'wb' to create a new file in write binary mode.
#3. Loop over the Response object’s iter_content() method.
#4. Call write() on each iteration to write the content to the file.
#5. Call close() to close the file.
############################
#Need some HTML basics before you can really pick apart web pages
#Everything is placed between starting and ending tags to separate out 
#the content into different roles/contexts
#html is plaintext with the .html file extension

#rightclick pages to see source files, hit f12 to look into dev view
#using regex's to find info in HTML is advised against - hard to do
#Beautiful soup is used to extract information from a html page
#module titled bs4
import requests, bs4
res = requests.get('http://nostarch.com')
res.raise_for_status()
noStarchSoup = bs4.BeautifulSoup(res.text)
type(noStarchSoup)
#or load an html file from your harddrive
import os
os.chdir('C:\\Users\\Matthew\\Documents\\GitHub\\BioInf')
exampleFile = open('example.html')
exampleSoup = bs4.BeautifulSoup(exampleFile)
type(exampleSoup)
#retreve web page element by calling select() and passing a CSS selector
#somilar to regex, find patterns in html
#table in page 246 of CSS selector examples
#VERY DEEP TOPIC TO GET CONFUSED BY
import bs4
exampleFile = open('example.html')
exampleSoup = bs4.BeautifulSoup(exampleFile.read())
elems = exampleSoup.select('#Author')
type(elems)
len(elems) #should be 1 according to the book
type(elems[0])
#already broken, no idea why
pElems = exampleSoup.select('p')
str(pElems[0])
pElems[0].getText()
str(pElems[1])
pElems[1].getText()
str(pElems[2])
pElems[2].getText()
#2 projects described,1 a google search and auto link open in new tab 
#2 a downloader for all XKCD comics
#no interest in these ends at the moment, can always return
#end chapter 11

#Start Chapter 12 - Working with Excel Spreadsheets w openpyxl
#one excel file is a 'workbook' that can contain multiple 'sheets'
#columns in letters, rows in numbers - each cell has a coordinate
import os
os.chdir('C:\\Users\\Matthew\\Documents\\GitHub\\BioInf')
import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
type(wb)
#Book is broken and outdated - great
from openpyxl import load_workbook
wb = load_workbook('example.xlsx')
print(wb.sheetnames)
sheet3 = wb['Sheet3']
sheet1 = wb['Sheet1']
sheet1['A1'].value
sheet1['B3'].value
c = sheet1['B1']
c.value
'Row ' + str(c.row) + ', Column ' + str(c.column) + ' is ' + c.value
#c.column is returning a integer in place of the letter.. hmm
#auto converting?
sheet1['C1'].value
c.column
#can specify specific locations by row and column number
sheet1.cell(row=1, column=2).value
for i in range(1, 8, 2): # 1-8, counting by 2's == Only odd #'s
    print(i, sheet1.cell(row=i, column=2).value)
#size of sheet commands  OLD AND OUTDATED
sheet1.get_highest_row()
sheet1.get_highest_column()
#now:
sheet1.max_row() #errors
sheet1.max_row #This works, no () for some reason
#got the answer - max_row is not a function its a value
sheet1.max_column
#rows and columns from sheets
tuple(sheet1['A1':'C3'])
for rowOfCellObjects in sheet1['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')

sheet1[1] #row 1 
for i in sheet1[1]:
    print(i.value)

sheet1.columns[1] #old and broken

list(sheet1.columns)[1] #new and not broken
for obj in list(sheet1.columns)[1]:
    print(obj.value)
#Example Project - State Data
wb = load_workbook('censuspopdata.xlsx')
print(wb.sheetnames)
sheet = wb['Population by Census Tract']
#now the sheet is called sheet and loaded in ready to go
countyData = {}
print('Reading rows...')
for row in range(2, sheet.max_row +1):
    state = sheet['B' +str(row)].value
    county = sheet['C' +str(row)].value
    pop = sheet['D' +str(row)].value
#now to open a new text file and write contents to it
#dictionary with state abbreviations as keys
    #makes sure key for this state exists
    countyData.setdefault(state, {})
    #makes sure key for county in state exists
    countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})
    #each row rep one census tract, so increment by 1
    countyData[state][county]['tracts'] += 1
    #increase the county pop by the pop in this census tract
    countyData[state][county]['pop'] += int(pop)
#all pop data will be now tabulated and keyed by state
#time to write to a new file
    print('Writing results..')
    resultFile = open('census2010.py', 'w')
    resultFile.write('allData = ' + pprint.pformat(countyData))
    resultFile.close()
    print('Done.')
import census2010
census2010.allData['AK']['Anchorage']




































